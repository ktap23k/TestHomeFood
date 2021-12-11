# This is a sample Python script.

# Press ⌃R to execute it or replace it with your code.
# Press Double ⇧ to search everywhere for classes, files, tool windows, actions, and settings.
import time
import openpyxl
from selenium import webdriver
from selenium.webdriver import ActionChains, Keys


class KTGIOHANG:
    log = ''

    def __init__(self):
        self.start()

    def start(self):
        self.driver = webdriver.Chrome('./Driver/chromedriver')
        self.driver.implicitly_wait(10)
        self.driver.maximize_window()
        self.driver.get('https://homefoodshop.vn/')
        self.driver.find_element_by_xpath('/html/body/div[2]/div/div[2]/div/div/div[1]/div/div[1]/div/div[1]/div').click()
        self.driver.find_element_by_xpath('/html/body/div[2]/div/div[3]/span/button').click()

    def exit_threaring(self):
        self.driver.close()

    def test1(self):
        '''Test Case 1'''
        try:
            self.driver.find_element_by_xpath('/html/body/div[1]/div/div/div[1]/nav/div[1]/div[3]/ul/li[2]/span/span/div/div[1]').click()
            # self.driver.get('https://homefoodshop.vn/cart')
            text = self.driver.find_element_by_xpath('/html/body/div[1]/div/div/div[2]/div[2]/div/div[3]/div/span').text
        except:
            text = ''
            self.log += '/TestCase1 Error/'
        if text:
            text = 'pass'
        return text

    def test2(self):
        '''Test Case 2'''
        try:
            self.driver.get('https://homefoodshop.vn/')
            self.driver.execute_script("window.scrollTo(0, 400);")
            # self.driver.find_element_by_xpath('/html/body/div[1]/div/div/div[2]/div[2]/div/div[1]/div[2]/div[2]/div/div[1]/div/div[2]/div/div/div[2]/div').click()
            button = self.driver.find_element_by_xpath(
                '/html/body/div/div/div/div[2]/div[2]/div/div[1]/div[2]/div[2]/div/div[1]/div/div[2]/div/div/div[2]/div')
            # time.sleep(2)
            self.driver.execute_script("arguments[0].click();", button)
            time.sleep(2)
            self.driver.get('https://homefoodshop.vn/cart')
            try:
                text = self.driver.find_element_by_xpath('/html/body/div[1]/div/div/div[2]/div[2]/div/div[3]/table/tbody/tr/td[1]/div').text
            except:
                text = ''
                self.log += '/TestCase2 Error/'
        except:
            text = ''
            self.log += '/TestCase2 Error/'
        if text:
            text = 'pass'
        return text

    def test3(self):
        try:
            self.driver.get('https://homefoodshop.vn/')
            self.driver.execute_script("window.scrollTo(0, 400);")
            # self.driver.find_element_by_xpath('/html/body/div[1]/div/div/div[2]/div[2]/div/div[1]/div[2]/div[2]/div/div[1]/div/div[2]/div/div/div[2]/div').click()
            button = self.driver.find_element_by_xpath(
                '/html/body/div/div/div/div[2]/div[2]/div/div[1]/div[2]/div[2]/div/div[2]/div/div[2]/div/div/div[2]/div')
            # time.sleep(2)
            self.driver.execute_script("arguments[0].click();", button)
            time.sleep(2)
            self.driver.get('https://homefoodshop.vn/cart')
            try:
                text = self.driver.find_element_by_xpath('/html/body/div[1]/div/div/div[2]/div[2]/div/div[3]/table/tbody/tr/td[1]/div').text
                text1 = self.driver.find_element_by_xpath('/html/body/div[1]/div/div/div[2]/div[2]/div/div[3]/table/tbody/tr[2]/td[1]/div').text
                text += 'sản phẩm 2:' + text1
            except:
                text = ''
                self.log += '/TestCase3 Error/'
        except:
            text = ''
            self.log += '/TestCase3 Error/'
        if text:
            text = 'pass'
        return text

    def test4(self):
        try:
            self.driver.get('https://homefoodshop.vn/')
            self.driver.execute_script("window.scrollTo(0, 400);")
            # self.driver.find_element_by_xpath('/html/body/div[1]/div/div/div[2]/div[2]/div/div[1]/div[2]/div[2]/div/div[1]/div/div[2]/div/div/div[2]/div').click()
            button = self.driver.find_element_by_xpath(
                '/html/body/div/div/div/div[2]/div[2]/div/div[1]/div[2]/div[2]/div/div[2]/div/div[2]/div/div/div[2]/div')
            # time.sleep(2)
            self.driver.execute_script("arguments[0].click();", button)
            time.sleep(2)
            self.driver.get('https://homefoodshop.vn/cart')
            try:
                # select = self.driver.find_element_by_xpath('/html/body/div[1]/div/div/div[2]/div[2]/div/div[3]/table/tbody/tr[2]/td[5]/div/div/div/input')
                select = self.driver.find_element_by_xpath('/html/body/div/div/div/div[2]/div[2]/div/div[3]/table/tbody/tr[2]/td[5]/div/div/div/input')
                text = select.get_attribute('aria-valuenow')
            except:
                text = ''
                self.log += '/TestCase4 Error/'
        except:
            text = ''
            self.log += '/TestCase4 Error/'
        if text:
            text = 'pass'
        return text

    def test5(self):
        try:
            text = self.test4()
        except:
            text =''
            self.log += '/TestCase5 Error/'
        return text

    def test6(self):
        try:
            select = self.driver.find_element_by_xpath('/html/body/div/div/div/div[2]/div[2]/div/div[3]/table/tbody/tr[2]/td[5]/div/div/div/input')
            text = int(select.get_attribute('aria-valuenow'))
            select = self.driver.find_element_by_xpath('/html/body/div/div/div/div[2]/div[2]/div/div[3]/table/tbody/tr[1]/td[5]/div/div/div/input')
            text += int(select.get_attribute('aria-valuenow'))
            kt = self.driver.find_element_by_xpath('/html/body/div/div/div/div[1]/nav/div[1]/div[3]/ul/li[2]/span/span/div/div[2]')
            if int(text) == int(kt.text):
                text = 'Ok'
        except:
            text =''
            self.log += '/TestCase6 Error/'

        if text:
            text = 'pass'
        return text

    def test7(self):
        try:
            select = self.driver.find_element_by_xpath('/html/body/div/div/div/div[2]/div[2]/div/div[3]/table/tbody/tr[1]/td[5]/div/div/div/input')
            text = int(select.get_attribute('aria-valuenow'))
            self.driver.find_element_by_xpath('/html/body/div/div/div/div[2]/div[2]/div/div[3]/table/tbody/tr[1]/td[5]/div/div/span[1]').click()
            select = self.driver.find_element_by_xpath('/html/body/div/div/div/div[2]/div[2]/div/div[3]/table/tbody/tr[1]/td[5]/div/div/div/input')
            text2 = int(select.get_attribute('aria-valuenow'))
            if text == text2:
                text = 'pass'
        except:
            text = ''
            self.log += '/TestCase7 Error/'

        if text:
            text = 'pass'
        return text

    def test8(self):
        try:
            select = self.driver.find_element_by_xpath('/html/body/div/div/div/div[2]/div[2]/div/div[3]/table/tbody/tr[2]/td[5]/div/div/div/input')
            text = int(select.get_attribute('aria-valuenow'))
            self.driver.find_element_by_xpath('/html/body/div/div/div/div[2]/div[2]/div/div[3]/table/tbody/tr[2]/td[5]/div/div/span[1]').click()
            select = self.driver.find_element_by_xpath('/html/body/div/div/div/div[2]/div[2]/div/div[3]/table/tbody/tr[2]/td[5]/div/div/div/input')
            text2 = int(select.get_attribute('aria-valuenow'))
            if int(text) == int(text2)-1:
                text = 'pass'
        except:
            text = ''
            self.log += '/TestCase8 Error/'

        if text:
            text = 'pass'
        return text

    def test9(self):
        try:
            select = self.driver.find_element_by_xpath('/html/body/div/div/div/div[2]/div[2]/div/div[3]/table/tbody/tr[2]/td[5]/div/div/div/input')
            text = int(select.get_attribute('aria-valuenow'))
            self.driver.find_element_by_xpath('/html/body/div/div/div/div[2]/div[2]/div/div[3]/table/tbody/tr[2]/td[5]/div/div/span[2]').click()
            select = self.driver.find_element_by_xpath('/html/body/div/div/div/div[2]/div[2]/div/div[3]/table/tbody/tr[2]/td[5]/div/div/div/input')
            text2 = int(select.get_attribute('aria-valuenow'))
            if int(text) == int(text2)-1:
                text = 'pass'
        except:
            text = ''
            self.log += '/TestCase9 Error/'

        if text:
            text = 'pass'
        return text

    def test10(self):
        try:
            self.driver.find_element_by_xpath('/html/body/div/div/div/div[2]/div[2]/div/div[3]/table/tbody/tr[2]/td[7]/div/button').click()
            self.driver.find_element_by_xpath('/html/body/div[2]/div/div[3]/button[2]').click()
            text = 'pass'
        except:
            text =''
            self.log += '/TestCase10 Error/'
        return text

    def test11(self):
        try:
            self.driver.find_element_by_xpath('/html/body/div[1]/div/div/div[2]/div[2]/div/div[3]/table/tbody/tr/td[7]/div/button').click()
            self.driver.find_element_by_xpath('/html/body/div[2]/div/div[3]/button[2]').click()
            text = self.test1()
        except:
            text = ''
            self.log += '/TestCase11 Error/'
        return text

    def test12(self, data):
        def check(self,k):
            text1 = self.driver.find_element_by_xpath('/html/body/div/div/div/div[2]/div[2]/div/div[3]/table/tbody/tr/td[6]/div').text
            inp = self.driver.find_element_by_xpath('/html/body/div/div/div/div[2]/div[2]/div/div[3]/table/tbody/tr/td[5]/div/div/div/input')
            inp.clear()
            inp.send_keys(k)
            inp.send_keys(Keys.RETURN)
            time.sleep(1)
            text2 = self.driver.find_element_by_xpath('/html/body/div/div/div/div[2]/div[2]/div/div[3]/table/tbody/tr/td[6]/div').text

            if text1 == text2:
                return 1
            return 0
        try:
            self.test2()
            # data[0] = am; 1: float; 2 so 0; 3 max
            if check(self,data[0]) and  check(self, data[1]) and check(self,data[2]) and check(self,data[3]):
                text = 'pass'
            text =''
        except:
            text = ''
            self.log += '/TestCase12 Error/'

        return text

class TIMKIEM:
    def __init__(self):
        self.start()

    def end_thread(self):
        self.driver.close()

    def start(self):
        self.driver = webdriver.Chrome('./Driver/chromedriver')
        self.driver.implicitly_wait(5)
        self.driver.maximize_window()
        self.driver.get('https://homefoodshop.vn/')
        self.driver.find_element_by_xpath('/html/body/div[2]/div/div[2]/div/div/div[1]/div/div[1]/div/div[1]/div').click()
        self.driver.find_element_by_xpath('/html/body/div[2]/div/div[3]/span/button').click()

    def test(self, text):
        find = self.driver.find_element_by_xpath('/html/body/div/div/div/div[1]/nav/div[1]/div[2]/div/input')
        try:
            find.send_keys(text)
            find.send_keys(Keys.RETURN)
            li = self.driver.find_element_by_xpath('/html/body/div/div/div/div[2]/div[2]/div/div[1]/div[2]/div[2]/div/div[1]')
            out = 'pass'
        except:
            out = 'fail'
        time.sleep(1)
        find.clear()
        return out

    def runTest(self):
        data = self.readData()
        print(data)
        for i,j in zip(data,range(len(data))):
            text = self.test(i)
            self.sheet['C'+str(j+2)] = text
            if text == 'pass':
                my_red = openpyxl.styles.colors.Color(rgb='669900')
                my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
                self.sheet['C'+str(j+2)].fill = my_fill
            else:
                my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
                self.sheet['C' + str(j + 2)].fill = my_fill
        self.wbin.save('Input.xlsx')
    def readData(self):
        self.wbin = openpyxl.load_workbook('Input.xlsx')
        self.sheet = self.wbin['Sheet2']
        data = []
        for i in range(2, 10):
            data.append(self.sheet['B' + str(i)].value)
            if data[i-2] ==None:
                data[i-2]=''
        return data

class DATHANG:
    log = ''

    def __init__(self):
        self.start()

    def end_theadring(self):
        self.driver.close()

    def start(self):
        self.driver = webdriver.Chrome('./Driver/chromedriver')
        self.driver.implicitly_wait(5)
        self.driver.maximize_window()
        self.driver.get('https://homefoodshop.vn/')
        self.driver.find_element_by_xpath(
            '/html/body/div[2]/div/div[2]/div/div/div[1]/div/div[1]/div/div[1]/div').click()
        self.driver.find_element_by_xpath('/html/body/div[2]/div/div[3]/span/button').click()

    def test1(self, data):
        ''' Data chứa phone, password và tên Khách Hàng'''
        try:
            self.driver.get('https://homefoodshop.vn/')
            self.driver.execute_script("window.scrollTo(0, 400);")
            # self.driver.find_element_by_xpath('/html/body/div[1]/div/div/div[2]/div[2]/div/div[1]/div[2]/div[2]/div/div[1]/div/div[2]/div/div/div[2]/div').click()
            button = self.driver.find_element_by_xpath(
                '/html/body/div/div/div/div[2]/div[2]/div/div[1]/div[2]/div[2]/div/div[1]/div/div[2]/div/div/div[2]/div')
            # time.sleep(2)
            self.driver.execute_script("arguments[0].click();", button)
            time.sleep(1)
            # login with phonenumber and password
            self.driver.find_element_by_xpath('/html/body/div/div/div/div[1]/nav/div[1]/div[3]/ul/li[1]/span').click()
            self.driver.find_element_by_xpath('/html/body/div[3]/div/div[2]/div/div[2]/div[1]/div/div[1]/form/div[1]/div/div/input').send_keys(data['phone'])
            self.driver.find_element_by_xpath('/html/body/div[3]/div/div[2]/div/div[2]/div[1]/div/div[1]/form/div[2]/div/div/input').send_keys(data['password'])
            time.sleep(1)
            self.driver.find_element_by_xpath('/html/body/div[3]/div/div[2]/div/div[2]/div[1]/div/div[1]/form/div[2]/div/div/input').send_keys(Keys.RETURN)
            time.sleep(1)
            self.driver.get('https://homefoodshop.vn/shipment-info')
            name = self.driver.find_element_by_xpath('/html/body/div/div/div/div[2]/div[2]/div[1]/div[1]/div[2]/span[2]').text
            phone = self.driver.find_element_by_xpath('/html/body/div/div/div/div[2]/div[2]/div[1]/div[1]/div[3]/span[2]').text
            if name == data['name'] and phone == data['phone']:
                text = 'pass'
            else:
                text = 'fail'
        except Exception as e:
            text = 'fail'
            self.log += '/TestCase1 Error/'
            # print(e)
        return text

    def test2(self, data):
        '''Điều kiện: run test1. Data chứa phone và tên Khách Hàng'''
        try:
            name = self.driver.find_element_by_xpath('/html/body/div/div/div/div[2]/div[2]/div[1]/div[2]/form/div[1]/div/div/input').get_attribute('value')
            phone = self.driver.find_element_by_xpath('/html/body/div/div/div/div[2]/div[2]/div[1]/div[2]/form/div[2]/div/div/input').get_attribute('value')
            if name == data['name'] and phone == data['phone']:
                text = 'pass'
            else:
                text = 'fail'
        except:
            text = 'fail'
            self.log += '/TestCase2 Error/'
        return text

    def test3(self):
        '''điều kiện: run test 1-2'''
        try:
            n = self.driver.find_element_by_xpath('/html/body/div/div/div/div[2]/div[2]/div[1]/div[2]/form/div[1]/div/div/input')
            while n.get_attribute('value'):
                n.send_keys(Keys.BACKSPACE)
            time.sleep(1)
            name = self.driver.find_element_by_xpath('/html/body/div/div/div/div[2]/div[2]/div[1]/div[2]/form/div[1]/div/div[2]').text

            k = self.driver.find_element_by_xpath('/html/body/div/div/div/div[2]/div[2]/div[1]/div[2]/form/div[2]/div/div/input')
            while k.get_attribute('value'):
                k.send_keys(Keys.BACKSPACE)
            n.click()
            time.sleep(1)
            phone = self.driver.find_element_by_xpath('/html/body/div/div/div/div[2]/div[2]/div[1]/div[2]/form/div[2]/div/div[2]').text
            if name =='Vui lòng nhập họ và tên' and phone =='Vui lòng nhập số điện thoại':
                text = 'pass'
            else:
                text = 'fail'
        except:
            text = 'fail'
            self.log += '/TestCase3 Error/'
        return text

    def test4(self, data):
        try:
            clickCheckBox = self.driver.find_element_by_xpath('/html/body/div/div/div/div[2]/div[2]/div[1]/div[2]/label/span[1]')
            clickCheckBox.click()
            clickCheckBox.click()
            text = self.test2(data)
        except:
            text = 'fail'
            self.log += '/TestCase4 Error/'
        return text

    def test5(self):
        try:
            addr = self.driver.find_element_by_xpath('/html/body/div/div/div/div[2]/div[2]/div[2]/div[1]/div[2]/div/label/span[2]').text
            button = self.driver.find_element_by_xpath('/html/body/div/div/div/div[2]/div[3]/button[2]')
            self.driver.execute_script("arguments[0].click();", button)
            time.sleep(2)
            addr2 = self.driver.find_element_by_xpath('/html/body/div/div/div/div[2]/div[5]/div[3]/span[2]').text
            if addr == addr2:
                text ='pass'
            else:
                text = 'fail'
        except:
            text='fail'
            self.log += '/TestCase5 Error/'
        return text

    def test6(self):
        try:
            k = self.driver.find_element_by_xpath('/html/body/div/div/div/div[2]/div[7]/div/ul/li/div/div[2]').text
            if k=='Thanh toán tiền mặt':
                text ='pass'
            else:
                text ='fail'
        except:
            text = 'fail'
            self.log += '/TestCase6 Error/'
        return text

    def test7(self):
        '''Không nên test vì ảnh hưởng tới shop'''
        try:
            button = self.driver.find_element_by_xpath('/html/body/div/div/div/div[2]/div[8]/button[2]')
            self.driver.execute_script("arguments[0].click();", button)
            time.sleep(2)
            self.driver.get('https://homefoodshop.vn/my-account/list-order')
            t = self.driver.find_element_by_xpath('/html/body/div[1]/div/div/div[2]/div/div[2]/div/div[4]/div/span').text
            if t !='Bạn chưa có đơn hàng nào' and t:
                text = 'pass'
            else:
                text = 'fail'
        except:
            text = 'fail'
            self.log += '/TestCase7 Error/'
        return text