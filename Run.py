from main import KTGIOHANG, TIMKIEM, DATHANG
import openpyxl


def output_Excel(input_detail, path):
    # Xác định số hàng và cột lớn nhất trong file excel cần tạo
    row = len(input_detail)
    column = len(input_detail[0])
    # Tạo một workbook mới và active nó
    wb = openpyxl.Workbook()
    ws = wb.active
    # Dùng vòng lặp for để ghi nội dung từ input_detail vào file Excel
    for i in range(0, row):
        for j in range(0, column):
            v = input_detail[i][j]
            ws.cell(column=j + 1, row=i + 1, value=v)
            if v == 'fail':
                my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
                my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
                ws.cell(column=j + 1, row=i + 1).fill = my_fill
            elif v == 'pass':
                my_red = openpyxl.styles.colors.Color(rgb='669900')
                my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
                ws.cell(column=j + 1, row=i + 1).fill = my_fill
    # Lưu lại file Excel
    wb.save(path)

def TestCaseGioHang():
    # create driver
    a = KTGIOHANG()
    wbin = openpyxl.load_workbook('Input.xlsx')
    sheet = wbin['Sheet1']
    data = []
    for i in range(2,6):
        data.append(sheet['A' + str(i)].value)
    print(data)

    inp = [['Số thứ tự', 'Tên Test Case', 'Output']]
    li = [a.test1(), a.test2(), a.test3(), a.test4(), a.test5(),a.test6(), a.test7(), a.test8(), a.test9(), a.test10(), a.test11(), a.test12(data)]
    for i in range(1,13):
        if li[i-1]:
            inp.append([i, 'TestCase ' +str(i), li[i-1]])
        else:
            inp.append([i, 'TestCase ' + str(i), 'fail'])
    output_Excel(inp,'./outputGioHang.xlsx')
    a.exit_threaring()
    print(a.log)

def TestCaseTimKiem():
    a = TIMKIEM()
    a.runTest()
    a.end_thread()

def TestCaseDatHang():
    a = DATHANG()
    wbin = openpyxl.load_workbook('Input.xlsx')
    sheet = wbin['Sheet3']
    data = {'name':'','phone':'', 'gmail': '', 'password':''}

    for i,j in zip(range(2, 6),data):
        data[j] = sheet['B' + str(i)].value

    li = [a.test1(data),a.test2(data), a.test3(), a.test4(data), a.test5(), a.test6(), 'fail']
    inp = [['Số thứ tự', 'Tên Test Case', 'Output']]
    for i in range(len(li)):
        if li[i]=='pass':
            inp.append([i+1, 'TestCase ' +str(i+1), li[i]])
        else:
            inp.append([i+1, 'TestCase ' + str(i+1), 'fail'])
    print(inp)
    output_Excel(inp,'./outputThanhToan.xlsx')
    #vì test case 7 sẽ đặt hàng ảnh hưởng tới shop, nhưng đặt hàng xong web k có đơn hàng đã đặt nên tự đặt test 7 thành fail sẵn, Đã test
    a.end_theadring()


if __name__ == '__main__':
    # TestCaseGioHang()
    # TestCaseTimKiem()
    TestCaseDatHang()